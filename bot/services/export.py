import io
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from bot.services.sheets import get_sheets_service


class ExportService:
    def __init__(self):
        self._sheets = None

    @property
    def sheets(self):
        if self._sheets is None:
            self._sheets = get_sheets_service()
        return self._sheets

    def generate_xlsx(self) -> bytes:
        records = self.sheets.get_all_records()
        columns = [
            ("ID", "ID"),
            ("Ngay luu", "Ngày lưu"),
            ("Tieu de", "Tiêu đề"),
            ("Link goc", "Link gốc"),
            ("Nguon", "Nguồn"),
            ("Tom tat AI", "Tóm tắt AI"),
            ("Ghi chu tay", "Ghi chú tay"),
            ("Chu de", "Chủ đề"),
            ("Tags", "Tags"),
            ("Uu tien", "Ưu tiên"),
            ("Trang thai", "Trạng thái"),
            ("Nguoi luu", "Người lưu"),
            ("Thumbnail", "Thumbnail"),
            ("Library Group", "Library Group"),
            ("Nhac nho", "Nhắc nhở"),
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "InfoSaver Data"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, (_, display_header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=display_header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        high_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        medium_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        low_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        applied_fill = PatternFill(start_color="D0E0E3", end_color="D0E0E3", fill_type="solid")

        for row_idx, record in enumerate(records, 2):
            for col_idx, (key, _) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

            priority = str(record.get("Uu tien", "")).strip().lower()
            status = str(record.get("Trang thai", "")).strip().lower()
            priority_cell = ws.cell(row=row_idx, column=10)
            status_cell = ws.cell(row=row_idx, column=11)
            if priority == "high":
                priority_cell.fill = high_fill
            elif priority == "medium":
                priority_cell.fill = medium_fill
            elif priority == "low":
                priority_cell.fill = low_fill

            if status == "da_ap_dung":
                status_cell.fill = applied_fill

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 25
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 15
        ws.column_dimensions["M"].width = 30
        ws.column_dimensions["N"].width = 18
        ws.column_dimensions["O"].width = 15

        ws.freeze_panes = "A2"
        last_column = get_column_letter(len(columns))
        last_row = max(2, len(records) + 1)
        ws.auto_filter.ref = f"A1:{last_column}{last_row}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
